import streamlit as st
import pandas as pd
from io import BytesIO, StringIO
import datetime
from pathlib import Path
from num2words import num2words
import xlrd
from openpyxl import load_workbook
from pdfrw import PdfReader, PdfWriter, PdfDict, PdfObject

st.set_page_config(page_title="Generador de Nota de Venta y LP", layout="wide")

# Logo CTC más grande arriba a la izquierda
st.markdown(
    """
    <div style='display: flex; align-items: center;'>
        <img src='https://raw.githubusercontent.com/marcossami/generador-LP/main/logo-ctcgroup.png' width='220' style='margin-right: 15px;'>
    </div>
    """,
    unsafe_allow_html=True
)

st.title("Generador Automático de Nota de Venta y Líquido Producto")

def cargar_consumos(raw_bytes, filename):
    ext = Path(filename).suffix.lower()
    buf = BytesIO(raw_bytes)
    if ext == ".csv":
        return pd.read_csv(StringIO(raw_bytes.decode("utf-8", errors="replace")))
    try:
        if ext == ".xls":
            return pd.read_excel(buf, engine="xlrd")
        return pd.read_excel(buf, engine="openpyxl")
    except:
        try:
            return pd.read_html(raw_bytes, header=0)[0]
        except:
            pass
    raise RuntimeError("Error leyendo consumos; convíertelo a .xlsx o .csv.")

# --- helper de formato AR: miles con punto y decimales con coma ---
def fmt_money(v):
    return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ─── Sidebar Inputs (solo cambian etiquetas) ───────────────────────────────────
reporte_file     = st.sidebar.file_uploader("Reporte De Consumos",     type=["xls","xlsx","csv"])
proveedores_file = st.sidebar.file_uploader("Listado De Proveedores",  type=["xls","xlsx","csv"])
plantilla_file   = st.sidebar.file_uploader("Plantilla Editable",      type=["pdf"])
numero_lp        = st.sidebar.text_input("Número de LP")

# IVA fijo al 21%
iva_rate = 21.0

# ─── Fechas ───────────────────────────────────────────────────────────────────
hoy     = datetime.date.today()
lp_date = datetime.date(hoy.year, hoy.month, 1)
lp_str  = lp_date.strftime("%d/%m/%Y")
if lp_date.month == 1:
    prev_m, prev_y = 12, lp_date.year - 1
else:
    prev_m, prev_y = lp_date.month - 1, lp_date.year
periodo_liq = f"{prev_m:02d}/{prev_y}"

# ─── Workflow ─────────────────────────────────────────────────────────────────
if reporte_file and proveedores_file and plantilla_file:
    # 1) Leer consumos y extraer Subtotal (I11), marca (B7) y B8 (KANSAS)
    raw_cons = reporte_file.read()
    try:
        df_cons   = cargar_consumos(raw_cons, reporte_file.name)
        ext_cons  = Path(reporte_file.name).suffix.lower()
        if ext_cons == ".xls":
            wb_xls   = xlrd.open_workbook(file_contents=raw_cons)
            sheet    = wb_xls.sheet_by_index(0)
            subtotal = float(sheet.cell_value(10, 8))  # I11
            marca    = str(sheet.cell_value(6, 1)).strip()  # B7
            try:
                b8_text = str(sheet.cell_value(7, 1)).strip()  # B8
            except:
                b8_text = ""
        else:
            wb_xlsx  = load_workbook(filename=BytesIO(raw_cons), data_only=True)
            ws       = wb_xlsx.active
            subtotal = float(ws["I11"].value)
            marca    = str(ws["B7"].value).strip()
            try:
                b8_text = str(ws["B8"].value).strip() if ws["B8"].value is not None else ""
            except:
                b8_text = ""
    except Exception as e:
        st.error(f"Error leyendo consumos: {e}")
        st.stop()

    # 2) Leer proveedores y preparar buscador
    raw_prov = proveedores_file.read()
    try:
        df_prov = cargar_consumos(raw_prov, proveedores_file.name)
    except Exception as e:
        st.error(f"Error leyendo proveedores: {e}")
        st.stop()

    # Limpiamos razón social y marcas
    df_prov["ProvClean"]   = df_prov["Proveedores"].str.replace(r"\s*\(.*\)", "", regex=True).str.strip()
    df_prov["MarcaClean"]  = df_prov["marcas"].astype(str).str.strip()

    # 3) Buscar proveedor (regla especial KANSAS con B8)
    if str(marca).strip().upper() == "KANSAS":
        ref_texto = "KANSAS - [MARTIN FIERRO 3361] - UDAONDO - BUENOS AIRES - 10010846"
        target_cuit = "30-71667294-4" if b8_text == ref_texto else "30-69765269-4"
        matches = df_prov[df_prov["CUIT"].astype(str).str.strip() == target_cuit]
    else:
        matches = df_prov[df_prov["MarcaClean"].str.lower() == marca.lower()]

    if matches.empty:
        st.error(f"No se encontró proveedor para la marca: '{marca}'")
        st.stop()
    prov = matches.iloc[0]

    # Mostrar marca y proveedor detectado
    st.write(f"🔍 Marca detectada en reporte: **{marca}**")
    st.write("Proveedor seleccionado automáticamente:")
    st.write(f"- Razón social: {prov['ProvClean']}")
    st.write(f"- Dirección:    {prov['Dirección']}")
    st.write(f"- CUIT:         {prov['CUIT']}")

    # 4) Generar LP
    if st.button("Generar LP"):
        try:
            tpl = PdfReader(BytesIO(plantilla_file.read()))

            # (Opcional) Mostrar campos detectados
            if tpl.Root.AcroForm and tpl.Root.AcroForm.Fields:
                campos = [f.T[1:-1] for f in tpl.Root.AcroForm.Fields if f.T]
                st.write("📑 Campos internos detectados:", campos)

            # Calcular IVA y Total y armar literal (enteros y decimales en letras)
            iva_amt = round(subtotal * iva_rate / 100, 2)
            total   = round(subtotal + iva_amt, 2)
            entero  = int(total)
            decs    = int(round((total - entero) * 100))

            entero_letras = num2words(entero, lang="es").capitalize()
            if decs > 0:
                decs_letras = num2words(decs, lang="es")
                literal = f"{entero_letras} con {decs_letras}"
            else:
                literal = entero_letras

            # Mapeo con formato AR en los montos
            valores = {
                "cliente":     prov["ProvClean"],    # Señor/es
                "dirección":   prov["Dirección"],   # Dirección
                "iva":         "Inscripto",         # I.V.A.
                "cuit":        prov["CUIT"],        # C.U.I.T.
                "fecha":       lp_str,              # Fecha
                "nfactura":    numero_lp,           # Nº LP
                "detalle":     "ventas por cuenta y orden",
                "subtotal":    fmt_money(subtotal),
                "iva insc":    fmt_money(iva_amt),
                "iva total":   fmt_money(total),
                "liquidacion": periodo_liq,
                "enpesos":     literal,             # Son Pesos
            }

            # Rellenar campos sin paréntesis
            for page in tpl.pages:
                if not page.Annots:
                    continue
                for annot in page.Annots:
                    if annot.T:
                        key = annot.T[1:-1]
                        if key in valores:
                            annot.V  = valores[key]
                            annot.AP = None

            # Forzar regenerar apariencias
            if tpl.Root.AcroForm:
                tpl.Root.AcroForm.update(PdfDict(NeedAppearances=PdfObject("true")))

            # Guardar y descargar
            out = BytesIO()
            PdfWriter().write(out, tpl)
            out.seek(0)
            st.download_button(
                "Descargar Nota de Venta y LP (PDF)",
                data=out,
                file_name="Nota_Venta_LP.pdf",
                mime="application/pdf"
            )
            st.success("¡LP generado correctamente!")
        except Exception as e:
            st.error(f"Error generando LP: {e}")
else:
    st.info(" Por favor, cargar reporte de consumos, listado de proveedores y plantilla PDF para comenzar.")

